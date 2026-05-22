"""
Informe: Estado de Lead por Fuente de Tráfico — Pipeline RST
Abril y Mayo diferenciados.

Fuente: hs_analytics_source (original). Si vacío → hs_latest_source (más reciente).
Misma lógica de estados que informe_rst_lead_status.py
"""
import requests
import pandas as pd
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()
console = Console()

TOKEN = os.getenv("HUBSPOT_TOKEN")
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
BASE = "https://api.hubapi.com"

RST_PIPELINE_ID = "1688043709"

RST_STAGES = {
    "2289712318": "Cita programada",
    "2289712320": "Entrevista realizada",
    "2289712322": "Envío de inscripción",
    "2289712323": "Cierre ganado",
    "2289712324": "Cierre perdido",
}

PERIODOS = {
    "Abril": ("2026-04-01T00:00:00Z", "2026-04-30T23:59:59Z"),
    "Mayo":  ("2026-05-01T00:00:00Z", "2026-05-31T23:59:59Z"),
}

CONTACT_PROPS = [
    "firstname", "lastname", "email",
    "hs_lead_status",
    "num_contacted_notes",
    "estado_de_lead_no_valido",
    "hs_analytics_source",
    "hs_analytics_source_data_1",
    "hs_latest_source",
    "hs_latest_source_data_1",
]

DEAL_PROPS = [
    "dealname", "dealstage", "createdate",
    "motivos_de_cierre_perdido_rst",
    "closed_lost_reason",
]

FUENTES_ES = {
    "ORGANIC_SEARCH":  "Búsqueda orgánica",
    "PAID_SEARCH":     "Búsqueda pagada",
    "EMAIL_MARKETING": "Email marketing",
    "SOCIAL_MEDIA":    "Redes sociales",
    "REFERRALS":       "Referencias",
    "OTHER_CAMPAIGNS": "Otras campañas",
    "DIRECT_TRAFFIC":  "Tráfico directo",
    "OFFLINE":         "Offline",
    "PAID_SOCIAL":     "Social pagado",
    "AI_REFERRALS":    "Referral IA",
}

LEAD_STATUS_NORM = {
    "new":                  "Nuevo",
    "nuevo":                "Nuevo",
    "open":                 "Nuevo",
    "in_progress":          "Intentando contactar",
    "intentando contactar": "Intentando contactar",
    "connected":            "Contactado",
    "contactado":           "Contactado",
    "bad timing":           "No interés",
    "no interés":           "No interés",
    "no interes":           "No interés",
    "unqualified":          "No válido",
    "no válido":            "No válido",
    "no valido":            "No válido",
    "matriculado":          "Matriculado",
    "negocio cerrado":      "Negocio Cerrado",
    "closed":               "Negocio Cerrado",
    "ilocalizado":          "Ilocalizado",
    "uncontacted":          "Ilocalizado",
}

ESTADOS_ORDEN = [
    "Nuevo", "Intentando contactar", "Contactado",
    "No interés", "No válido", "Ilocalizado",
    "Matriculado", "Negocio Cerrado", "Sin estado",
]

COLORES_ESTADO = {
    "Nuevo":               "C8E6C9",
    "Intentando contactar":"FFF9C4",
    "Contactado":          "B3E5FC",
    "No interés":          "FFE0B2",
    "No válido":           "FFCDD2",
    "Ilocalizado":         "F3E5F5",
    "Matriculado":         "80CBC4",
    "Negocio Cerrado":     "FFCA28",
    "Sin estado":          "EEEEEE",
}

COLORES_FUENTE = {
    "Búsqueda orgánica": "E8F5E9",
    "Búsqueda pagada":   "E3F2FD",
    "Email marketing":   "FFF3E0",
    "Redes sociales":    "FCE4EC",
    "Social pagado":     "F3E5F5",
    "Tráfico directo":   "E0F7FA",
    "Otras campañas":    "FFF8E1",
    "Offline":           "EFEBE9",
    "Referencias":       "F1F8E9",
    "Referral IA":       "EDE7F6",
    "Sin datos":         "FAFAFA",
}


def resolve_fuente(cp):
    """Original si tiene valor, si no → más reciente. Siempre con etiqueta de origen."""
    raw_orig = (cp.get("hs_analytics_source") or "").strip()
    raw_rec  = (cp.get("hs_latest_source") or "").strip()

    if raw_orig:
        fuente = FUENTES_ES.get(raw_orig, raw_orig.replace("_", " ").title())
        origen = "Original"
        detalle = (cp.get("hs_analytics_source_data_1") or "").strip()
    elif raw_rec:
        fuente = FUENTES_ES.get(raw_rec, raw_rec.replace("_", " ").title())
        origen = "Más reciente"
        detalle = (cp.get("hs_latest_source_data_1") or "").strip()
    else:
        fuente = "Sin datos"
        origen = "—"
        detalle = ""

    return fuente, origen, detalle


def norm_lead_status(raw):
    if not raw:
        return "Sin estado"
    return LEAD_STATUS_NORM.get(raw.lower().strip(), raw.strip().title())


# ── Fetch ─────────────────────────────────────────────────────────────────────

def get_rst_deals(mes, inicio, fin):
    deals = []
    after = None
    with Progress(SpinnerColumn(), TextColumn(f"  Deals {mes}..."),
                  BarColumn(), TaskProgressColumn(), console=console) as p:
        task = p.add_task(mes, total=None)
        while True:
            payload = {
                "filterGroups": [{"filters": [
                    {"propertyName": "pipeline",   "operator": "EQ",  "value": RST_PIPELINE_ID},
                    {"propertyName": "createdate", "operator": "GTE", "value": inicio},
                    {"propertyName": "createdate", "operator": "LTE", "value": fin},
                ]}],
                "properties": DEAL_PROPS,
                "limit": 100,
            }
            if after:
                payload["after"] = after
            r = requests.post(f"{BASE}/crm/v3/objects/deals/search",
                              headers=HEADERS, json=payload)
            data = r.json()
            deals.extend(data.get("results", []))
            p.update(task, completed=len(deals))
            paging = data.get("paging", {})
            if not paging or not paging.get("next"):
                break
            after = paging["next"]["after"]
    return deals


def get_contact_ids(deal_ids):
    deal_to_contact = {}
    with Progress(SpinnerColumn(), TextColumn("  Asociaciones..."),
                  BarColumn(), TaskProgressColumn(), console=console) as p:
        task = p.add_task("assoc", total=len(deal_ids))
        for i in range(0, len(deal_ids), 100):
            batch = deal_ids[i:i+100]
            r = requests.post(
                f"{BASE}/crm/v4/associations/deals/contacts/batch/read",
                headers=HEADERS,
                json={"inputs": [{"id": d} for d in batch]},
            )
            if r.status_code == 200:
                for item in r.json().get("results", []):
                    fid = str(item.get("from", {}).get("id", ""))
                    tos = item.get("to", [])
                    if tos:
                        deal_to_contact[fid] = str(tos[0]["toObjectId"])
            p.update(task, completed=min(i+100, len(deal_ids)))
    return deal_to_contact


def get_contacts(contact_ids):
    contacts = {}
    unique = list(set(contact_ids))
    with Progress(SpinnerColumn(), TextColumn("  Contactos..."),
                  BarColumn(), TaskProgressColumn(), console=console) as p:
        task = p.add_task("contacts", total=len(unique))
        for i in range(0, len(unique), 100):
            batch = unique[i:i+100]
            r = requests.post(f"{BASE}/crm/v3/objects/contacts/batch/read",
                              headers=HEADERS,
                              json={"inputs": [{"id": c} for c in batch],
                                    "properties": CONTACT_PROPS})
            if r.status_code == 200:
                for c in r.json().get("results", []):
                    contacts[str(c["id"])] = c["properties"]
            p.update(task, completed=min(i+100, len(unique)))
    return contacts


def build_df(deals, deal_to_contact, contacts, mes):
    rows = []
    for deal in deals:
        did = str(deal["id"])
        dp = deal["properties"]
        etapa = RST_STAGES.get(dp.get("dealstage", ""), dp.get("dealstage", ""))
        motivo_cierre = dp.get("motivos_de_cierre_perdido_rst") or dp.get("closed_lost_reason") or ""

        cid = deal_to_contact.get(did)
        cp = contacts.get(cid, {}) if cid else {}

        fuente, origen_fuente, fuente_detalle = resolve_fuente(cp)
        lead_status = norm_lead_status(cp.get("hs_lead_status") or "")
        intentos = int(cp.get("num_contacted_notes") or 0)
        motivo_no_valido = cp.get("estado_de_lead_no_valido") or ""

        rows.append({
            "mes":              mes,
            "deal_id":          did,
            "fuente":           fuente,
            "origen_fuente":    origen_fuente,
            "fuente_detalle":   fuente_detalle,
            "lead_status":      lead_status,
            "intentos":         intentos,
            "motivo_no_valido": motivo_no_valido,
            "motivo_cierre":    motivo_cierre,
            "etapa_deal":       etapa,
        })
    return pd.DataFrame(rows)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def style_header(ws, color="1F4E79"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 38)
    ws.freeze_panes = "B2"


def color_fuente_rows(ws, fuente_col_idx):
    for row in ws.iter_rows(min_row=2):
        fuente = str(row[fuente_col_idx].value or "")
        hex_col = COLORES_FUENTE.get(fuente, "FFFFFF")
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=hex_col)
            cell.alignment = Alignment(horizontal="center")


# ── Hojas de informe ──────────────────────────────────────────────────────────

def hoja_resumen(df, writer, mes, color):
    """Pivot principal: Fuente × Estado de Lead."""
    dm = df[df["mes"] == mes]
    total = len(dm)

    pivot = (dm.groupby(["fuente", "lead_status"])
               .size()
               .reset_index(name="Total"))
    pivot_wide = pivot.pivot_table(index="fuente", columns="lead_status",
                                   values="Total", fill_value=0, aggfunc="sum")
    for estado in ESTADOS_ORDEN:
        if estado not in pivot_wide.columns:
            pivot_wide[estado] = 0
    pivot_wide = pivot_wide[[e for e in ESTADOS_ORDEN if e in pivot_wide.columns]]
    pivot_wide["TOTAL"] = pivot_wide.sum(axis=1)
    pivot_wide["% del Total"] = (pivot_wide["TOTAL"] / total * 100).round(1)
    pivot_wide = pivot_wide.sort_values("TOTAL", ascending=False).reset_index()
    pivot_wide.rename(columns={"fuente": "Fuente de Tráfico"}, inplace=True)

    sname = f"{mes} - Resumen por Fuente"
    pivot_wide.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)

    # Colorear columnas de estados
    header_map = {cell.value: cell.column for cell in ws[1]}
    for estado, hex_col in COLORES_ESTADO.items():
        if estado in header_map:
            col_idx = header_map[estado]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value and cell.value > 0:
                        cell.fill = PatternFill("solid", fgColor=hex_col)
                    cell.alignment = Alignment(horizontal="center")

    # Colorear columna fuente
    f_col = header_map.get("Fuente de Tráfico")
    if f_col:
        for row in ws.iter_rows(min_row=2, min_col=f_col, max_col=f_col):
            for cell in row:
                fuente = str(cell.value or "")
                cell.fill = PatternFill("solid", fgColor=COLORES_FUENTE.get(fuente, "FFFFFF"))
                cell.font = Font(bold=True)


def hoja_origen_fuente(df, writer, mes, color):
    """Desglose de cuántos vienen de fuente original vs más reciente."""
    dm = df[df["mes"] == mes]
    resumen = (dm.groupby(["fuente", "origen_fuente"])
                 .size().reset_index(name="Total"))
    pivot = resumen.pivot_table(index="fuente", columns="origen_fuente",
                                values="Total", fill_value=0, aggfunc="sum")
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False).reset_index()
    pivot.rename(columns={"fuente": "Fuente"}, inplace=True)

    sname = f"{mes} - Origen Fuente"
    pivot.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)

    header_map = {cell.value: cell.column for cell in ws[1]}
    f_col = header_map.get("Fuente")
    for row in ws.iter_rows(min_row=2):
        fuente = str(row[0].value or "")
        bg = COLORES_FUENTE.get(fuente, "FFFFFF")
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")


def hoja_intentos(df, writer, mes, color):
    """Intentando contactar: fuente, nº contactos, intentos totales y promedio."""
    dm = df[(df["mes"] == mes) & (df["lead_status"] == "Intentando contactar")]
    if dm.empty:
        return
    resumen = (dm.groupby("fuente")
                 .agg(Contactos=("deal_id","count"),
                      Total_intentos=("intentos","sum"),
                      Promedio_intentos=("intentos","mean"),
                      Max_intentos=("intentos","max"))
                 .round(1).sort_values("Contactos", ascending=False).reset_index())
    resumen.columns = ["Fuente", "Nº Contactos", "Total Intentos",
                        "Promedio Intentos", "Máx Intentos"]
    sname = f"{mes} - Intentando Contactar"[:31]
    resumen.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)
    fill = PatternFill("solid", fgColor=COLORES_ESTADO["Intentando contactar"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def hoja_no_valido(df, writer, mes, color):
    """No válido: fuente × motivo de invalidación."""
    dm = df[(df["mes"] == mes) & (df["lead_status"] == "No válido")]
    if dm.empty:
        return
    dm = dm.copy()
    dm["motivo_no_valido"] = dm["motivo_no_valido"].replace("", "Sin motivo especificado")
    pivot = (dm.groupby(["fuente", "motivo_no_valido"])
               .size().reset_index(name="Total"))
    pivot_wide = pivot.pivot_table(index="fuente", columns="motivo_no_valido",
                                   values="Total", fill_value=0, aggfunc="sum")
    pivot_wide["TOTAL"] = pivot_wide.sum(axis=1)
    pivot_wide = pivot_wide.sort_values("TOTAL", ascending=False).reset_index()
    pivot_wide.rename(columns={"fuente": "Fuente"}, inplace=True)

    sname = f"{mes} - No Válido (Motivos)"[:31]
    pivot_wide.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)
    fill = PatternFill("solid", fgColor=COLORES_ESTADO["No válido"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def hoja_cierre_perdido(df, writer, mes, color):
    """Negocio Cerrado: fuente × motivo de cierre perdido RST."""
    dm = df[(df["mes"] == mes) & (df["lead_status"] == "Negocio Cerrado")]
    if dm.empty:
        return
    dm = dm.copy()
    dm["motivo_cierre"] = dm["motivo_cierre"].replace("", "Sin motivo registrado")
    pivot = (dm.groupby(["fuente", "motivo_cierre"])
               .size().reset_index(name="Total"))
    pivot_wide = pivot.pivot_table(index="fuente", columns="motivo_cierre",
                                   values="Total", fill_value=0, aggfunc="sum")
    pivot_wide["TOTAL"] = pivot_wide.sum(axis=1)
    pivot_wide = pivot_wide.sort_values("TOTAL", ascending=False).reset_index()
    pivot_wide.rename(columns={"fuente": "Fuente"}, inplace=True)

    sname = f"{mes} - Neg. Cerrado (Motivos)"[:31]
    pivot_wide.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)
    fill = PatternFill("solid", fgColor=COLORES_ESTADO["Negocio Cerrado"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def hoja_detalle(df, writer, mes, color):
    """Detalle completo con fuente, origen y estado."""
    dm = df[df["mes"] == mes][[
        "fuente", "origen_fuente", "fuente_detalle", "lead_status",
        "intentos", "motivo_no_valido", "motivo_cierre", "etapa_deal",
    ]].sort_values(["fuente", "lead_status"])

    sname = f"{mes} - Detalle Completo"
    dm.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)

    header_map = {cell.value: cell.column for cell in ws[1]}
    status_col = header_map.get("lead_status")
    for row in ws.iter_rows(min_row=2):
        estado = str(row[status_col - 1].value or "") if status_col else ""
        hex_col = COLORES_ESTADO.get(estado, "FFFFFF")
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=hex_col)
            cell.alignment = Alignment(horizontal="left")


def hoja_comparativa(df, writer):
    """Comparativa Abril vs Mayo por fuente y estado."""
    comp = (df.groupby(["mes", "fuente", "lead_status"])
              .size().reset_index(name="Total"))
    pivot = comp.pivot_table(index=["fuente", "lead_status"], columns="mes",
                              values="Total", fill_value=0, aggfunc="sum").reset_index()
    for m in ["Abril", "Mayo"]:
        if m not in pivot.columns:
            pivot[m] = 0
    pivot["Variación"] = pivot["Mayo"] - pivot["Abril"]
    pivot["Var %"] = ((pivot["Variación"] / pivot["Abril"].replace(0, 1)) * 100).round(1)
    pivot.rename(columns={"fuente": "Fuente", "lead_status": "Estado Lead"}, inplace=True)
    pivot = pivot.sort_values(["Fuente", "Estado Lead"])

    sname = "COMPARATIVA Abr vs May"
    pivot.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, "4A148C")
    autofit(ws)

    header_map = {cell.value: cell.column for cell in ws[1]}
    var_col = header_map.get("Variación")
    if var_col:
        for row in ws.iter_rows(min_row=2):
            val = row[var_col - 1].value
            if val and val > 0:
                row[var_col - 1].fill = PatternFill("solid", fgColor="C8E6C9")
            elif val and val < 0:
                row[var_col - 1].fill = PatternFill("solid", fgColor="FFCDD2")


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    console.print("\n[bold blue]══════════════════════════════════════════════════[/bold blue]")
    console.print("[bold blue]  INFORME FUENTE DE TRÁFICO × ESTADO — RST ABR/MAY [/bold blue]")
    console.print("[bold blue]══════════════════════════════════════════════════[/bold blue]\n")

    colores = {"Abril": "1565C0", "Mayo": "2E7D32"}
    all_dfs = []

    for mes, (inicio, fin) in PERIODOS.items():
        console.print(f"\n[bold cyan]── {mes.upper()} ──[/bold cyan]")
        deals = get_rst_deals(mes, inicio, fin)
        if not deals:
            console.print(f"[yellow]Sin deals en {mes}[/yellow]")
            continue

        deal_to_contact = get_contact_ids([d["id"] for d in deals])
        contacts = get_contacts(list(set(deal_to_contact.values())))
        dm = build_df(deals, deal_to_contact, contacts, mes)
        all_dfs.append(dm)

        # Preview
        dist_fuente = dm["fuente"].value_counts()
        console.print(f"  Fuentes {mes}: " +
                      " | ".join(f"{k}: {v}" for k, v in dist_fuente.items()))
        n_reciente = (dm["origen_fuente"] == "Más reciente").sum()
        n_original = (dm["origen_fuente"] == "Original").sum()
        n_sin = (dm["origen_fuente"] == "—").sum()
        console.print(f"  Origen fuente → Original: {n_original} | "
                      f"Más reciente (fallback): {n_reciente} | Sin datos: {n_sin}")

    if not all_dfs:
        console.print("[red]Sin datos.[/red]")
        return

    df = pd.concat(all_dfs, ignore_index=True)
    os.makedirs("exports", exist_ok=True)
    path = "exports/informe_rst_fuente_trafico.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for mes in ["Abril", "Mayo"]:
            if mes not in df["mes"].values:
                continue
            color = colores[mes]
            hoja_resumen(df, writer, mes, color)
            hoja_origen_fuente(df, writer, mes, color)
            hoja_intentos(df, writer, mes, color)
            hoja_no_valido(df, writer, mes, color)
            hoja_cierre_perdido(df, writer, mes, color)
            hoja_detalle(df, writer, mes, color)

        hoja_comparativa(df, writer)

        df.to_excel(writer, sheet_name="DATOS COMPLETOS", index=False)
        style_header(writer.sheets["DATOS COMPLETOS"], "37474F")
        autofit(writer.sheets["DATOS COMPLETOS"])

    console.print(f"\n[bold green]Informe generado: {path}[/bold green]")
    console.print(f"[dim]Total: {len(df)} registros "
                  f"(Abril: {len(df[df['mes']=='Abril'])} | "
                  f"Mayo: {len(df[df['mes']=='Mayo'])})[/dim]")
    return path


if __name__ == "__main__":
    run()
